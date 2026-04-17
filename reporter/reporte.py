#!/usr/bin/env python3
"""
Reporte Semanal de Performance — Mosca Hnos.
Obtiene datos de Meta Ads, Google Ads, TikTok Ads y GA4.
Genera un email HTML y lo envía a agustina@cardinal.com.uy

Horario: lunes 12:00 (semana anterior) | viernes 14:00 (semana en curso)
"""

import os, json, smtplib, sys
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from dotenv import load_dotenv
import requests

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

# ── CREDENCIALES ──────────────────────────────────────────────────────────────

META_ACCESS_TOKEN    = os.getenv("META_ACCESS_TOKEN")
META_AD_ACCOUNT_ID   = os.getenv("META_AD_ACCOUNT_ID")       # "act_XXXXXXXXXX"
META_TOKEN_DATE      = os.getenv("META_TOKEN_DATE")           # fecha de generación del token

GOOGLE_DEVELOPER_TOKEN = os.getenv("GOOGLE_DEVELOPER_TOKEN")
GOOGLE_CLIENT_ID       = os.getenv("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET   = os.getenv("GOOGLE_CLIENT_SECRET")
GOOGLE_REFRESH_TOKEN   = os.getenv("GOOGLE_REFRESH_TOKEN")
GOOGLE_CUSTOMER_ID     = os.getenv("GOOGLE_CUSTOMER_ID")     # cuenta Mosca Hnos.
GOOGLE_MCC_ID          = os.getenv("GOOGLE_MCC_ID")           # cuenta administradora

TIKTOK_ACCESS_TOKEN  = os.getenv("TIKTOK_ACCESS_TOKEN")
TIKTOK_ADVERTISER_ID = os.getenv("TIKTOK_ADVERTISER_ID")

GA4_PROPERTY_ID = os.getenv("GA4_PROPERTY_ID")               # solo el número
GA4_KEY_FILE    = os.getenv("GA4_KEY_FILE", os.path.join(os.path.dirname(__file__), "ga4-key.json"))

EMAIL_SENDER       = os.getenv("EMAIL_SENDER")                # agustina@cardinal.com.uy
EMAIL_APP_PASSWORD = os.getenv("EMAIL_APP_PASSWORD")          # App Password de Google
EMAIL_RECIPIENT    = os.getenv("EMAIL_RECIPIENT", "agustina@cardinal.com.uy")

ANTHROPIC_API_KEY  = os.getenv("ANTHROPIC_API_KEY")           # para análisis IA en Notas

# ── FECHAS ────────────────────────────────────────────────────────────────────

def check_meta_token_expiry():
    """Devuelve un mensaje de alerta si el token de Meta vence en menos de 10 días."""
    if not META_TOKEN_DATE or META_TOKEN_DATE == "never":
        return None
    try:
        token_date  = datetime.strptime(META_TOKEN_DATE, "%Y-%m-%d")
        expiry_date = token_date + timedelta(days=60)
        days_left   = (expiry_date - datetime.today()).days
        if days_left <= 10:
            return (f"⚠️ <strong>El token de Meta Ads vence en {days_left} días "
                    f"({expiry_date.strftime('%d/%m/%Y')}).</strong> "
                    f"Renovalo en developers.facebook.com/tools/explorer "
                    f"y actualizá META_ACCESS_TOKEN en el archivo .env.")
    except Exception:
        pass
    return None


def get_date_range():
    """
    Desde el 1° del mes actual hasta ayer.
    """
    today     = datetime.today()
    yesterday = today - timedelta(days=1)
    start     = today.replace(day=1)
    return start.strftime("%Y-%m-%d"), yesterday.strftime("%Y-%m-%d")


# ── META ADS ──────────────────────────────────────────────────────────────────

def get_meta_campaigns(start_date, end_date):
    # Paso 1: objetivos reales por campaña (opcional — falla si el token no tiene acceso)
    try:
        camp_r = requests.get(
            f"https://graph.facebook.com/v19.0/{META_AD_ACCOUNT_ID}/campaigns",
            params={"fields": "id,objective", "limit": 200, "access_token": META_ACCESS_TOKEN},
            timeout=30,
        )
        camp_r.raise_for_status()
        objectives = {c["id"]: c.get("objective", "—") for c in camp_r.json().get("data", [])}
    except Exception:
        objectives = {}

    PURCHASE_TYPES = {"offsite_conversion.fb_pixel_purchase", "purchase", "omni_purchase"}

    def _meta_pur(item):
        return int(float(next(
            (a["value"] for a in item.get("actions", []) if a["action_type"] in PURCHASE_TYPES), 0
        )))

    def _meta_valor(item):
        return round(float(next(
            (a["value"] for a in item.get("action_values", []) if a["action_type"] in PURCHASE_TYPES), 0.0
        )), 2)

    def _meta_cpa(item):
        pur = _meta_pur(item)
        inv = float(item.get("spend", 0))
        return round(inv / pur, 2) if pur > 0 else 0.0

    # Paso 2: insights con campaign_id para poder unir
    url = f"https://graph.facebook.com/v19.0/{META_AD_ACCOUNT_ID}/insights"
    params = {
        "level":       "campaign",
        "fields":      "campaign_id,campaign_name,spend,reach,impressions,"
                       "cpm,frequency,clicks,cpc,ctr,actions,action_values,cost_per_action_type",
        "time_range":  json.dumps({"since": start_date, "until": end_date}),
        "access_token": META_ACCESS_TOKEN,
        "limit":       100,
    }
    r = requests.get(url, params=params, timeout=30)
    if not r.ok:
        print("Meta error body:", r.text[:800])
    r.raise_for_status()

    rows = []
    for item in r.json().get("data", []):
        obj = objectives.get(item.get("campaign_id", ""), "—")
        rows.append({
            "plataforma":    "Meta",
            "fuente_medio":  "facebook / cpc",
            "objetivo":      obj,
            "campana":       item.get("campaign_name", "—"),
            "inversion":     float(item.get("spend", 0)),
            "alcance":       int(item.get("reach", 0)),
            "impresiones":   int(item.get("impressions", 0)),
            "cpm":           float(item.get("cpm", 0)),
            "frecuencia":    float(item.get("frequency", 0)),
            "clicks":        int(item.get("clicks", 0)),
            "cpc":           float(item.get("cpc", 0)),
            "compras_meta":  _meta_pur(item),
            "cpa_meta":      _meta_cpa(item),
            "valor_meta":    _meta_valor(item),
            "ctr":           float(item.get("ctr", 0)),
        })
    return sorted(rows, key=lambda x: (x["objetivo"], -x["inversion"]))


# ── GOOGLE ADS ────────────────────────────────────────────────────────────────

def get_google_campaigns(start_date, end_date):
    from google.ads.googleads.client import GoogleAdsClient

    client = GoogleAdsClient.load_from_dict({
        "developer_token":   GOOGLE_DEVELOPER_TOKEN,
        "client_id":         GOOGLE_CLIENT_ID,
        "client_secret":     GOOGLE_CLIENT_SECRET,
        "refresh_token":     GOOGLE_REFRESH_TOKEN,
        "login_customer_id": GOOGLE_MCC_ID.replace("-", ""),
        "use_proto_plus":    True,
    })

    customer_id = GOOGLE_CUSTOMER_ID.replace("-", "")
    ga_service  = client.get_service("GoogleAdsService")

    query = f"""
        SELECT
            campaign.name,
            campaign.advertising_channel_type,
            metrics.cost_micros,
            metrics.impressions,
            metrics.clicks,
            metrics.ctr,
            metrics.average_cpc
        FROM campaign
        WHERE campaign.status = 'ENABLED'
          AND metrics.impressions > 0
          AND segments.date BETWEEN '{start_date}' AND '{end_date}'
        ORDER BY metrics.cost_micros DESC
    """

    channel_labels = {
        "SEARCH":          ("google / cpc",      "Search"),
        "SHOPPING":        ("google / shopping",  "Shopping"),
        "PERFORMANCE_MAX": ("google / pmax",      "Performance Max"),
        "DISPLAY":         ("google / display",   "Display"),
        "VIDEO":           ("google / video",     "Video"),
    }

    rows = []
    stream = ga_service.search_stream(customer_id=customer_id, query=query)
    for batch in stream:
        for result in batch.results:
            c  = result.campaign
            m  = result.metrics
            ch = c.advertising_channel_type.name
            fuente, objetivo = channel_labels.get(ch, ("google / cpc", ch))

            cost   = m.cost_micros / 1_000_000
            impr   = m.impressions
            clicks = m.clicks
            ctr    = m.ctr * 100
            cpc    = m.average_cpc / 1_000_000
            cpm    = (cost / impr * 1000) if impr > 0 else 0

            rows.append({
                "plataforma":   "Google",
                "fuente_medio": fuente,
                "objetivo":     objetivo,
                "campana":      c.name,
                "inversion":    round(cost, 2),
                "alcance":      "—",
                "impresiones":  impr,
                "cpm":          round(cpm, 2),
                "frecuencia":   "—",
                "clicks":       clicks,
                "cpc":          round(cpc, 2),
                "ctr":          round(ctr, 2),
            })
    return sorted(rows, key=lambda x: (x["objetivo"], -x["inversion"]))


# ── TIKTOK ADS ────────────────────────────────────────────────────────────────

def get_tiktok_campaigns(start_date, end_date):
    url     = "https://business-api.tiktok.com/open_api/v1.3/report/integrated/get/"
    headers = {"Access-Token": TIKTOK_ACCESS_TOKEN}
    params  = {
        "advertiser_id": TIKTOK_ADVERTISER_ID,
        "report_type":   "BASIC",
        "data_level":    "AUCTION_CAMPAIGN",
        "dimensions":    json.dumps(["campaign_id"]),
        "metrics":       json.dumps([
            "campaign_name", "objective_type",
            "spend", "reach", "impressions", "cpm",
            "frequency", "clicks", "cpc", "ctr",
        ]),
        "start_date":    start_date,
        "end_date":      end_date,
        "page_size":     100,
    }

    r = requests.get(url, headers=headers, params=params, timeout=30)
    r.raise_for_status()

    rows = []
    for item in r.json().get("data", {}).get("list", []):
        m = item.get("metrics", {})
        rows.append({
            "plataforma":   "TikTok",
            "fuente_medio": "tiktok / paid",
            "objetivo":     m.get("objective_type", "—"),
            "campana":      m.get("campaign_name", "—"),
            "inversion":    float(m.get("spend", 0)),
            "alcance":      int(m.get("reach", 0)),
            "impresiones":  int(m.get("impressions", 0)),
            "cpm":          float(m.get("cpm", 0)),
            "frecuencia":   float(m.get("frequency", 0)),
            "clicks":       int(m.get("clicks", 0)),
            "cpc":          float(m.get("cpc", 0)),
            "ctr":          float(m.get("ctr", 0)),
        })
    return sorted(rows, key=lambda x: x["inversion"], reverse=True)


# ── GOOGLE ANALYTICS 4 ────────────────────────────────────────────────────────

def get_ga4_data(start_date, end_date):
    from google.analytics.data_v1beta import BetaAnalyticsDataClient
    from google.analytics.data_v1beta.types import (
        RunReportRequest, Dimension, Metric, DateRange,
    )
    from google.oauth2 import service_account

    creds  = service_account.Credentials.from_service_account_file(
        GA4_KEY_FILE,
        scopes=["https://www.googleapis.com/auth/analytics.readonly"],
    )
    client = BetaAnalyticsDataClient(credentials=creds)

    req = RunReportRequest(
        property=f"properties/{GA4_PROPERTY_ID}",
        dimensions=[Dimension(name="sessionCampaignName")],
        metrics=[
            Metric(name="totalUsers"),
            Metric(name="sessions"),
            Metric(name="addToCarts"),
            Metric(name="checkouts"),
            Metric(name="ecommercePurchases"),
            Metric(name="purchaseRevenue"),
        ],
        date_ranges=[DateRange(start_date=start_date, end_date=end_date)],
    )

    response = client.run_report(req)

    data = {}
    for row in response.rows:
        source = row.dimension_values[0].value
        data[source] = {
            "usuarios":       int(row.metric_values[0].value),
            "sesiones":       int(row.metric_values[1].value),
            "add_to_cart":    int(row.metric_values[2].value),
            "begin_checkout": int(row.metric_values[3].value),
            "purchase":       int(row.metric_values[4].value),
            "revenue":        float(row.metric_values[5].value),
        }
    return data


def get_ga4_daily_by_channel(start_date, end_date):
    """GA4 diario por canal — para separar Meta (Paid Social) y Google (Paid Search/Shopping)."""
    from google.analytics.data_v1beta import BetaAnalyticsDataClient
    from google.analytics.data_v1beta.types import RunReportRequest, Dimension, Metric, DateRange
    from google.oauth2 import service_account

    creds  = service_account.Credentials.from_service_account_file(
        GA4_KEY_FILE,
        scopes=["https://www.googleapis.com/auth/analytics.readonly"],
    )
    client = BetaAnalyticsDataClient(credentials=creds)

    req = RunReportRequest(
        property=f"properties/{GA4_PROPERTY_ID}",
        dimensions=[
            Dimension(name="date"),
            Dimension(name="sessionDefaultChannelGroup"),
        ],
        metrics=[
            Metric(name="sessions"),
            Metric(name="addToCarts"),
            Metric(name="checkouts"),
            Metric(name="ecommercePurchases"),
            Metric(name="purchaseRevenue"),
        ],
        date_ranges=[DateRange(start_date=start_date, end_date=end_date)],
    )

    response = client.run_report(req)
    data = {}
    for row in response.rows:
        raw   = row.dimension_values[0].value
        fecha = f"{raw[:4]}-{raw[4:6]}-{raw[6:]}"
        ch    = row.dimension_values[1].value
        if fecha not in data:
            data[fecha] = {}
        data[fecha][ch] = {
            "sesiones":       int(row.metric_values[0].value),
            "add_to_cart":    int(row.metric_values[1].value),
            "begin_checkout": int(row.metric_values[2].value),
            "purchase":       int(row.metric_values[3].value),
            "revenue":        float(row.metric_values[4].value),
        }
    return data


def get_ga4_daily_totals(start_date, end_date):
    """Totales del sitio por día sin dimensión de campaña."""
    from google.analytics.data_v1beta import BetaAnalyticsDataClient
    from google.analytics.data_v1beta.types import RunReportRequest, Dimension, Metric, DateRange
    from google.oauth2 import service_account

    creds  = service_account.Credentials.from_service_account_file(
        GA4_KEY_FILE,
        scopes=["https://www.googleapis.com/auth/analytics.readonly"],
    )
    client = BetaAnalyticsDataClient(credentials=creds)

    req = RunReportRequest(
        property=f"properties/{GA4_PROPERTY_ID}",
        dimensions=[Dimension(name="date")],
        metrics=[
            Metric(name="sessions"),
            Metric(name="addToCarts"),
            Metric(name="checkouts"),
            Metric(name="ecommercePurchases"),
            Metric(name="purchaseRevenue"),
        ],
        date_ranges=[DateRange(start_date=start_date, end_date=end_date)],
    )

    response = client.run_report(req)
    data = {}
    for row in response.rows:
        # GA4 devuelve fecha como YYYYMMDD → convertir a YYYY-MM-DD
        raw = row.dimension_values[0].value
        fecha = f"{raw[:4]}-{raw[4:6]}-{raw[6:]}"
        data[fecha] = {
            "sesiones":       int(row.metric_values[0].value),
            "add_to_cart":    int(row.metric_values[1].value),
            "begin_checkout": int(row.metric_values[2].value),
            "purchase":       int(row.metric_values[3].value),
            "revenue":        float(row.metric_values[4].value),
        }
    return data


# ── EMAIL HTML ────────────────────────────────────────────────────────────────

PLAT_COLOR = {"Meta": "#1877F2", "Google": "#4285F4", "TikTok": "#010101"}

def _d(v, fmt="num"):
    if v == "—" or v is None: return "—"
    if fmt == "currency": return f"${float(v):,.2f}"
    if fmt == "pct":      return f"{float(v):.2f}%"
    if fmt == "freq":     return f"{float(v):.2f}x"
    return f"{int(v):,}"


def _fmt(date_str):
    """Convierte YYYY-MM-DD a DD/MM/YYYY para mostrar."""
    d = datetime.strptime(date_str, "%Y-%m-%d")
    return d.strftime("%d/%m/%Y")

def _logo_base64():
    import base64
    logo_path = os.path.join(os.path.dirname(__file__), "logo-mosca.png")
    try:
        with open(logo_path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")
    except Exception:
        return None

def build_email_html(start_date, end_date, meta, google, tiktok, ga4):
    all_rows  = meta + google + tiktok
    total_inv = sum(r["inversion"] for r in all_rows)
    total_ses = sum(v.get("sesiones", 0) for v in ga4.values())
    total_pur = sum(v.get("purchase", 0) for v in ga4.values())
    total_rev = sum(v.get("revenue", 0.0) for v in ga4.values())
    cpa_prom  = (total_inv / total_pur) if total_pur > 0 else 0
    conv_total = (total_pur / total_ses * 100) if total_ses > 0 else 0

    weekday  = datetime.today().weekday()
    periodo  = "Semana anterior" if weekday == 0 else "Semana en curso"
    gen_date = datetime.today().strftime("%d/%m/%Y %H:%M")

    # ── GA4 lookup por nombre de campaña (exacto o parcial) ──────────────────
    def ga4_by_campaign(campana):
        name = campana.lower()
        if name in ga4:
            return ga4[name]
        for key, val in ga4.items():
            if name in key.lower() or key.lower() in name:
                return val
        return {}

    # ── Table rows ────────────────────────────────────────────────────────────
    def make_rows(rows):
        html = ""
        for i, r in enumerate(rows):
            bg    = "#f9f9f9" if i % 2 == 0 else "#ffffff"
            color = PLAT_COLOR.get(r["plataforma"], "#333")
            td    = f'padding:7px 10px;border-bottom:1px solid #eee;background:{bg};font-size:12px;'
            campana = str(r["campana"])
            campana_short = (campana[:38] + "…") if len(campana) > 38 else campana
            ga    = ga4_by_campaign(campana)
            rev      = ga.get("revenue", 0.0)
            pur      = ga.get("purchase", 0)
            ses      = ga.get("sesiones", 0)
            inv      = r["inversion"]
            roas     = (rev / inv) if inv > 0 else 0
            conv     = (pur / ses * 100) if ses > 0 else 0
            html += f"""<tr>
              <td style="{td}color:{color};font-weight:600;" title="{campana}">{campana_short}</td>
              <td style="{td}text-align:right;">{_d(r['inversion'],'currency')}</td>
              <td style="{td}text-align:right;">{_d(r['alcance'])}</td>
              <td style="{td}text-align:right;">{_d(r['impresiones'])}</td>
              <td style="{td}text-align:right;">{_d(r['cpm'],'currency')}</td>
              <td style="{td}text-align:right;">{_d(r['frecuencia'],'freq')}</td>
              <td style="{td}text-align:right;">{_d(r['clicks'])}</td>
              <td style="{td}text-align:right;">{_d(r['cpc'],'currency')}</td>
              <td style="{td}text-align:right;">{_d(r['ctr'],'pct')}</td>
              <td style="{td}text-align:right;">{_d(ga.get('usuarios'))}</td>
              <td style="{td}text-align:right;">{_d(ga.get('sesiones'))}</td>
              <td style="{td}text-align:right;">{_d(ga.get('add_to_cart'))}</td>
              <td style="{td}text-align:right;">{_d(ga.get('begin_checkout'))}</td>
              <td style="{td}text-align:right;">{_d(ga.get('purchase'))}</td>
              <td style="{td}text-align:right;">{_d(rev,'currency') if rev else '—'}</td>
              <td style="{td}text-align:right;">{f'{roas:.2f}x' if roas else '—'}</td>
              <td style="{td}text-align:right;font-weight:700;">{f'{conv:.2f}%' if conv else '—'}</td>
            </tr>"""
        return html

    def platform_totals(rows):
        inv     = sum(r["inversion"] for r in rows)
        alcance = sum(r["alcance"] for r in rows if isinstance(r["alcance"], int))
        impr    = sum(r["impresiones"] for r in rows if isinstance(r["impresiones"], int))
        clicks  = sum(r["clicks"] for r in rows if isinstance(r["clicks"], int))
        cpc     = (inv / clicks) if clicks > 0 else 0
        # GA4 totals per platform
        g_users = g_ses = g_cart = g_checkout = g_pur = 0
        g_rev   = 0.0
        for r in rows:
            gd = ga4_by_campaign(r["campana"])
            g_users    += gd.get("usuarios", 0)
            g_ses      += gd.get("sesiones", 0)
            g_cart     += gd.get("add_to_cart", 0)
            g_checkout += gd.get("begin_checkout", 0)
            g_pur      += gd.get("purchase", 0)
            g_rev      += gd.get("revenue", 0.0)
        roas = (g_rev / inv) if inv > 0 else 0
        conv = (g_pur / g_ses * 100) if g_ses > 0 else 0
        return inv, alcance, impr, clicks, cpc, g_users, g_ses, g_cart, g_checkout, g_pur, g_rev, roas, conv

    def section_header(label, color, rows):
        count = len(rows)
        inv, alcance, impr, clicks, cpc, g_users, g_ses, g_cart, g_checkout, g_pur, g_rev, roas, conv = platform_totals(rows)
        hs = "background:{c};color:white;padding:8px 10px;font-size:11px;font-weight:700;".format(c=color)
        hn = "background:{c};color:white;padding:8px 10px;font-size:11px;font-weight:700;text-align:right;".format(c=color)
        alcance_cell = f'<td style="{hn}">{alcance:,}</td>' if alcance > 0 else f'<td style="{hn}">—</td>'
        return f"""<tr>
          <td style="{hs}text-transform:uppercase;letter-spacing:0.5px;">{label} ADS &nbsp;·&nbsp; {count} campañas</td>
          <td style="{hn}">${inv:,.2f}</td>
          {alcance_cell}
          <td style="{hn}">{impr:,}</td>
          <td style="{hn}">—</td>
          <td style="{hn}">—</td>
          <td style="{hn}">{clicks:,}</td>
          <td style="{hn}">${cpc:,.2f}</td>
          <td style="{hn}">—</td>
          <td style="{hn}">{g_users:,}</td>
          <td style="{hn}">{g_ses:,}</td>
          <td style="{hn}">{g_cart:,}</td>
          <td style="{hn}">{g_checkout:,}</td>
          <td style="{hn}">{g_pur:,}</td>
          <td style="{hn}">${g_rev:,.2f}</td>
          <td style="{hn}">{f'{roas:.2f}x' if roas else '—'}</td>
          <td style="{hn}">{f'{conv:.2f}%' if conv else '—'}</td>
        </tr>"""

    th_style = "padding:9px 10px;background:#00215E;color:white;font-size:11px;white-space:nowrap;text-align:left;"
    headers  = ["Campaña","Inversión","Alcance",
                 "Impr.","CPM","Frec.","Clicks","CPC","CTR",
                 "Usuarios GA4","Sesiones","Add Cart","Checkout","Compras",
                 "Ingresos","ROAS","Conv. %"]

    ths = "".join(f'<th style="{th_style}">{h}</th>' for h in headers)

    table = f"""
    <table style="border-collapse:collapse;width:100%;font-family:Arial,sans-serif;">
      <thead><tr>{ths}</tr></thead>
      <tbody>
        {section_header("Meta", PLAT_COLOR["Meta"], meta)}
        {make_rows(meta)}
        {section_header("Google", PLAT_COLOR["Google"], google)}
        {make_rows(google)}
        {section_header("TikTok", PLAT_COLOR["TikTok"], tiktok)}
        {make_rows(tiktok)}
      </tbody>
    </table>"""


    # ── Summary cards ─────────────────────────────────────────────────────────
    def card(label, value):
        return f"""
        <td style="padding:16px 24px;text-align:center;border-right:1px solid #e0e0e0;">
          <p style="margin:0;font-size:11px;color:#888;text-transform:uppercase;letter-spacing:1px;">{label}</p>
          <p style="margin:6px 0 0;font-size:22px;font-weight:700;color:#00215E;">{value}</p>
        </td>"""

    return f"""<!DOCTYPE html>
<html>
<body style="margin:0;padding:20px;background:#f0f2f5;font-family:Arial,sans-serif;">
<div style="max-width:1300px;margin:0 auto;background:#fff;border-radius:10px;
            overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.08);">

  <!-- Header -->
  <div style="background:#00215E;padding:28px 36px;">
    <table style="width:100%;border-collapse:collapse;">
      <tr>
        <td>
          <p style="margin:0;color:white;font-size:20px;font-weight:700;">Reporte de Performance</p>
          <p style="margin:5px 0 0;color:rgba(255,255,255,0.7);font-size:13px;">
            Mosca Hnos. &nbsp;·&nbsp; {periodo}
          </p>
        </td>
        <td style="text-align:right;vertical-align:middle;">
          <img src="cid:logo_mosca" style="max-height:48px;filter:brightness(0) invert(1);display:block;margin:0 0 8px auto;" />
          <p style="margin:0;color:rgba(255,255,255,0.9);font-size:13px;">{_fmt(start_date)} — {_fmt(end_date)}</p>
          <p style="margin:4px 0 0;color:rgba(255,255,255,0.5);font-size:11px;">Generado {gen_date}</p>
        </td>
      </tr>
    </table>
  </div>

  <!-- Summary -->
  <table style="width:100%;border-collapse:collapse;border-bottom:1px solid #eee;">
    <tr>
      {card("Inversión Total", f"${total_inv:,.2f}")}
      {card("Sesiones Totales", f"{total_ses:,}")}
      {card("Compras Totales", f"{total_pur:,}")}
      {card("CPA Promedio", f"${cpa_prom:,.2f}" if total_pur > 0 else "—")}
      {card("Ingresos Totales", f"${total_rev:,.2f}")}
      {card("Tasa de Conversión", f"{conv_total:.2f}%" if conv_total > 0 else "—")}
    </tr>
  </table>

  <!-- Table -->
  <div style="padding:28px 36px;overflow-x:auto;">
    {table}
  </div>

  {f'''<!-- Token Alert -->
  <div style="padding:14px 36px;background:#fff8e1;border-top:3px solid #E85D1A;">
    <p style="margin:0;font-size:13px;color:#c0392b;">{check_meta_token_expiry()}</p>
  </div>''' if check_meta_token_expiry() else ''}

  <!-- Footer -->
  <div style="padding:16px 36px;background:#f8f9fa;border-top:1px solid #eee;">
    <p style="margin:0;font-size:11px;color:#bbb;">
      Generado automáticamente · Cardinal Agency ·
      Datos: Meta Ads API · Google Ads API · TikTok Marketing API · GA4 Data API
    </p>
  </div>

</div>
</body>
</html>"""


# ── DATOS DIARIOS PARA EXCEL ──────────────────────────────────────────────────

def get_meta_daily(start_date, end_date):
    PURCHASE_TYPES = {"offsite_conversion.fb_pixel_purchase", "purchase", "omni_purchase"}

    def _pur(item):
        return int(float(next(
            (a["value"] for a in item.get("actions", []) if a["action_type"] in PURCHASE_TYPES), 0
        )))

    def _valor(item):
        return round(float(next(
            (a["value"] for a in item.get("action_values", []) if a["action_type"] in PURCHASE_TYPES), 0.0
        )), 2)

    url = f"https://graph.facebook.com/v19.0/{META_AD_ACCOUNT_ID}/insights"
    params = {
        "level":          "campaign",
        "fields":         "date_start,campaign_name,spend,reach,impressions,"
                          "cpm,frequency,clicks,cpc,ctr,actions,action_values",
        "time_range":     json.dumps({"since": start_date, "until": end_date}),
        "time_increment": 1,
        "access_token":   META_ACCESS_TOKEN,
        "limit":          500,
    }
    r = requests.get(url, params=params, timeout=30)
    r.raise_for_status()

    rows = []
    for item in r.json().get("data", []):
        inv = float(item.get("spend", 0))
        pur = _pur(item)
        rows.append({
            "fecha":        item.get("date_start", ""),
            "campana":      item.get("campaign_name", "—"),
            "inversion":    inv,
            "alcance":      int(item.get("reach", 0)),
            "impresiones":  int(item.get("impressions", 0)),
            "cpm":          float(item.get("cpm", 0)),
            "frecuencia":   float(item.get("frequency", 0)),
            "clicks":       int(item.get("clicks", 0)),
            "cpc":          float(item.get("cpc", 0)),
            "compras_meta": pur,
            "cpa_meta":     round(inv / pur, 2) if pur > 0 else 0.0,
            "valor_meta":   _valor(item),
            "ctr":          float(item.get("ctr", 0)),
        })
    return sorted(rows, key=lambda x: (x["campana"], x["fecha"]))


def get_google_daily(start_date, end_date):
    from google.ads.googleads.client import GoogleAdsClient

    client = GoogleAdsClient.load_from_dict({
        "developer_token":   GOOGLE_DEVELOPER_TOKEN,
        "client_id":         GOOGLE_CLIENT_ID,
        "client_secret":     GOOGLE_CLIENT_SECRET,
        "refresh_token":     GOOGLE_REFRESH_TOKEN,
        "login_customer_id": GOOGLE_MCC_ID.replace("-", ""),
        "use_proto_plus":    True,
    })

    customer_id = GOOGLE_CUSTOMER_ID.replace("-", "")
    ga_service  = client.get_service("GoogleAdsService")

    query = f"""
        SELECT
            segments.date,
            campaign.name,
            campaign.advertising_channel_type,
            metrics.cost_micros,
            metrics.impressions,
            metrics.clicks,
            metrics.ctr,
            metrics.average_cpc
        FROM campaign
        WHERE campaign.status = 'ENABLED'
          AND metrics.impressions > 0
          AND segments.date BETWEEN '{start_date}' AND '{end_date}'
        ORDER BY segments.date, metrics.cost_micros DESC
    """

    channel_labels = {
        "SEARCH":          "Search",
        "SHOPPING":        "Shopping",
        "PERFORMANCE_MAX": "Performance Max",
        "DISPLAY":         "Display",
        "VIDEO":           "Video",
    }

    rows = []
    stream = ga_service.search_stream(customer_id=customer_id, query=query)
    for batch in stream:
        for result in batch.results:
            c    = result.campaign
            m    = result.metrics
            s    = result.segments
            ch   = c.advertising_channel_type.name
            cost = m.cost_micros / 1_000_000
            impr = m.impressions
            rows.append({
                "fecha":       s.date,
                "campana":     c.name,
                "tipo":        channel_labels.get(ch, ch),
                "inversion":   round(cost, 2),
                "impresiones": impr,
                "cpm":         round((cost / impr * 1000) if impr > 0 else 0, 2),
                "clicks":      m.clicks,
                "cpc":         round(m.average_cpc / 1_000_000, 2),
                "ctr":         round(m.ctr * 100, 2),
            })
    return rows


# ── EXCEL EVOLUTIVO ───────────────────────────────────────────────────────────

def update_excel(start_date, end_date, meta, google, tiktok, ga4, meta_daily, google_daily, ga4_totals, ga4_by_channel, ga4_totals_yoy=None, inv_yoy=0.0, update_sheets=True, add_status_tab=True):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from collections import defaultdict

    month_str  = datetime.strptime(start_date, "%Y-%m-%d").strftime("%Y-%m")
    excel_path = os.path.join(os.path.dirname(__file__), f"Mosca | Performance | Reporte mensual {month_str}.xlsx")

    # GA4 lookup igual que en el email
    def ga4_by_campaign(campana):
        name = campana.lower()
        if name in ga4:
            return ga4[name]
        for key, val in ga4.items():
            if name in key.lower() or key.lower() in name:
                return val
        return {}

    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    BLUE      = "00215E"
    HF        = "Arial"
    COL_FONT  = Font(bold=True, color="FFFFFF", size=10, name=HF)
    COL_FILL  = PatternFill("solid", fgColor=BLUE)
    COL_ALIGN = Alignment(horizontal="center", vertical="center")
    logo_path = os.path.join(os.path.dirname(__file__), "logo-mosca.png")

    def ensure_sheet(wb, name, headers):
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            ncols = len(headers)

            # ── Fila 1: encabezado azul con logo ─────────────────────────────
            ws.row_dimensions[1].height = 48
            last_col_letter = get_column_letter(ncols)
            ws.merge_cells(f"A1:{last_col_letter}1")
            hdr_cell = ws["A1"]
            hdr_cell.fill      = PatternFill("solid", fgColor=BLUE)
            hdr_cell.font      = Font(bold=True, color="FFFFFF", size=14, name=HF)
            hdr_cell.alignment = Alignment(horizontal="right", vertical="center",
                                           indent=1)
            hdr_cell.value = "MOSCA HNOS.  "

            if os.path.exists(logo_path):
                img        = XLImage(logo_path)
                img.width  = 44
                img.height = 44
                ws.add_image(img, "A1")

            # ── Fila 2: separador vacío ───────────────────────────────────────
            ws.row_dimensions[2].height = 6
            for col in range(1, ncols + 1):
                ws.cell(row=2, column=col).fill = PatternFill("solid", fgColor=BLUE)

            # ── Fila 3: cabeceras de columnas ─────────────────────────────────
            ws.row_dimensions[3].height = 20
            for col, h in enumerate(headers, 1):
                cell            = ws.cell(row=3, column=col, value=h)
                cell.font       = COL_FONT
                cell.fill       = COL_FILL
                cell.alignment  = COL_ALIGN
                ws.column_dimensions[get_column_letter(col)].width = max(14, len(h) + 2)

        return wb[name]

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    fecha_ej = datetime.today().strftime("%Y-%m-%d %H:%M")
    all_rows = meta + google + tiktok

    USD = '"$"#,##0.00'
    PCT = '0.00'

    def fmt_row(ws, row_idx, money_cols=(), pct_cols=()):
        for col in money_cols:
            ws.cell(row=row_idx, column=col).number_format = USD
        for col in pct_cols:
            ws.cell(row=row_idx, column=col).number_format = PCT

    import re as _re
    _date_pat = _re.compile(r'^\d{4}-\d{2}-\d{2}$')
    def _sheet_dates(ws, min_row=4):
        """Devuelve el set de fechas YYYY-MM-DD ya presentes en col A."""
        return {
            str(ws.cell(row=r, column=1).value)
            for r in range(min_row, ws.max_row + 1)
            if _date_pat.match(str(ws.cell(row=r, column=1).value or ''))
        }

    # ── Hoja Resumen (una fila por día) ──────────────────────────────────────
    from openpyxl.chart import LineChart, PieChart, Reference

    RESUMEN_HEADERS = [
        "Fecha", "Sesiones", "Add Cart", "Checkout",
        "Compras", "Inversión", "Ingresos", "ROAS", "Conv %",
    ]
    # col 6=Inversión (USD), col 7=Ingresos (USD), col 9=Conv% (PCT)
    RESUMEN_MONEY = (6, 7)
    RESUMEN_PCT   = (9,)


    # Helper: fila resumen destacada (fondo azul claro)
    SUMM_FILL = PatternFill("solid", fgColor="D6E4F0")
    SUMM_FONT = Font(bold=True, size=10, name=HF)

    def append_summary_row(ws, row_data, money_cols=(), pct_cols=()):
        ws.append(row_data)
        row_idx = ws.max_row
        for col in range(1, len(row_data) + 1):
            cell      = ws.cell(row=row_idx, column=col)
            cell.fill = SUMM_FILL
            cell.font = SUMM_FONT
        fmt_row(ws, row_idx, money_cols=money_cols, pct_cols=pct_cols)

    if update_sheets:
        resumen_nueva = "Resumen" not in wb.sheetnames
        ws_r = ensure_sheet(wb, "Resumen", RESUMEN_HEADERS)
        # Mantener layout homogéneo en la tabla principal (A:I)
        resumen_widths = {1: 16, 2: 14, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14, 9: 14}
        for col, w in resumen_widths.items():
            ws_r.column_dimensions[get_column_letter(col)].width = w
        ws_r.row_dimensions[1].height = 48
        ws_r.row_dimensions[2].height = 6
        ws_r.row_dimensions[3].height = 20

        # Eliminar restos del embudo histórico y dejar Resumen solo en A:I
        to_unmerge = [str(mg) for mg in list(ws_r.merged_cells.ranges) if mg.max_col > 9]
        for mg in to_unmerge:
            ws_r.unmerge_cells(mg)
        if ws_r.max_column > 9:
            ws_r.delete_cols(10, ws_r.max_column - 9)

        # Inversión diaria (Meta + Google) para calcular ROAS por día
        inv_por_dia = defaultdict(float)
        for r in meta_daily:
            inv_por_dia[r["fecha"]] += r["inversion"]
        for r in google_daily:
            inv_por_dia[r["fecha"]] += r["inversion"]
        # ── Limpiar duplicados: eliminar filas de resumen que quedaron
        #    mezcladas entre los datos diarios (runs anteriores con bug) ──────
        rows_to_delete = []
        for rr in range(7, ws_r.max_row + 1):
            v = str(ws_r.cell(row=rr, column=1).value or "")
            if v in ("Año anterior", "% vs año anterior"):
                rows_to_delete.append(rr)
            elif ws_r.cell(row=rr, column=1).value is None:
                # fila con col A vacía pero con datos (resumen duplicado)
                if any(ws_r.cell(row=rr, column=c).value is not None for c in range(2, 10)):
                    rows_to_delete.append(rr)
        for rr in reversed(rows_to_delete):
            ws_r.delete_rows(rr)

        # ── Totales Analytics del período (siempre mismo rango analizado) ─────
        tot_ses  = sum(d.get("sesiones", 0)       for d in ga4_totals.values())
        tot_cart = sum(d.get("add_to_cart", 0)    for d in ga4_totals.values())
        tot_chk  = sum(d.get("begin_checkout", 0) for d in ga4_totals.values())
        tot_pur  = sum(d.get("purchase", 0)       for d in ga4_totals.values())
        tot_rev  = sum(d.get("revenue", 0.0)      for d in ga4_totals.values())
        tot_inv  = sum(inv_por_dia.values())
        tot_conv = round(tot_pur / tot_ses * 100, 2) if tot_ses > 0 else None

        # ── Totales Analytics YoY del mismo período ───────────────────────────
        has_yoy = bool(ga4_totals_yoy)
        yoy_ses = yoy_cart = yoy_chk = yoy_pur = yoy_rev = yoy_conv = None
        if has_yoy:
            yoy_ses  = sum(d.get("sesiones", 0)       for d in ga4_totals_yoy.values())
            yoy_cart = sum(d.get("add_to_cart", 0)    for d in ga4_totals_yoy.values())
            yoy_chk  = sum(d.get("begin_checkout", 0) for d in ga4_totals_yoy.values())
            yoy_pur  = sum(d.get("purchase", 0)       for d in ga4_totals_yoy.values())
            yoy_rev  = sum(d.get("revenue", 0.0)      for d in ga4_totals_yoy.values())
            yoy_conv = round(yoy_pur / yoy_ses * 100, 2) if yoy_ses > 0 else None

        YOY_FILL  = PatternFill("solid", fgColor="EAF0FB")
        YOY_FONT  = Font(size=10, name=HF, italic=True)
        DIFF_FILL = PatternFill("solid", fgColor="FFF3CD")
        DIFF_FONT = Font(bold=True, size=10, name=HF)

        def _write_summary_fixed_row(row_idx, row_data, fill, font, money_cols=(), pct_cols=()):
            for col, val in enumerate(row_data, 1):
                ws_r.cell(row=row_idx, column=col).value = val
            for col in range(1, len(row_data) + 1):
                c = ws_r.cell(row=row_idx, column=col)
                c.fill = fill
                c.font = font
                c.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")
            fmt_row(ws_r, row_idx, money_cols=money_cols, pct_cols=pct_cols)

        # Fila 4: Año anterior (si hay datos YoY)
        if has_yoy:
            _write_summary_fixed_row(4, [
                "Año anterior",
                yoy_ses  or None,
                yoy_cart or None,
                yoy_chk  or None,
                yoy_pur  or None,
                inv_yoy  or None,
                yoy_rev  or None,
                None,
                yoy_conv,
            ], YOY_FILL, YOY_FONT, money_cols=RESUMEN_MONEY, pct_cols=RESUMEN_PCT)

        # Fila 5: Período actual
        _write_summary_fixed_row(5, [
            None,
            tot_ses  or None,
            tot_cart or None,
            tot_chk  or None,
            tot_pur  or None,
            tot_inv  or None,
            tot_rev  or None,
            round(tot_rev / tot_inv, 2) if tot_inv > 0 else None,
            tot_conv,
        ], SUMM_FILL, SUMM_FONT, money_cols=RESUMEN_MONEY, pct_cols=RESUMEN_PCT)

        # Fila 6: % vs año anterior
        if has_yoy:
            def _pct_diff(curr, prev):
                if prev and prev != 0:
                    return round((curr - prev) / abs(prev) * 100, 1)
                return None
            _write_summary_fixed_row(6, [
                "% vs año anterior",
                _pct_diff(tot_ses,  yoy_ses),
                _pct_diff(tot_cart, yoy_cart),
                _pct_diff(tot_chk,  yoy_chk),
                _pct_diff(tot_pur,  yoy_pur),
                _pct_diff(tot_inv,  inv_yoy) if inv_yoy else None,
                _pct_diff(tot_rev,  yoy_rev),
                None,
                _pct_diff(tot_conv, yoy_conv) if (tot_conv and yoy_conv) else None,
            ], DIFF_FILL, DIFF_FONT)
            for col in (2, 3, 4, 5, 6, 9):
                ws_r.cell(row=6, column=col).number_format = '0.0\"%\"'
        else:
            for col in range(1, 10):
                ws_r.cell(row=6, column=col).value = None

        # ── Filas diarias Resumen (con dedup) ─────────────────────────────────
        existing_r = _sheet_dates(ws_r, min_row=7)
        last_data_row = 6
        for rr in range(7, ws_r.max_row + 1):
            if _date_pat.match(str(ws_r.cell(row=rr, column=1).value or "")):
                last_data_row = rr
        next_row = last_data_row + 1

        for fecha in sorted(ga4_totals):
            if fecha in existing_r:
                continue
            d   = ga4_totals[fecha]
            ses = d.get("sesiones", 0)
            pur = d.get("purchase", 0)
            rev = d.get("revenue", 0.0)
            inv = inv_por_dia.get(fecha, 0)
            row_data = [
                fecha,
                ses   or None,
                d.get("add_to_cart")    or None,
                d.get("begin_checkout") or None,
                pur   or None,
                inv   or None,
                rev   or None,
                round(rev / inv, 2) if inv > 0 else None,
                round(pur / ses * 100, 2) if ses > 0 else None,
            ]
            for col, val in enumerate(row_data, 1):
                ws_r.cell(row=next_row, column=col).value = val
            fmt_row(ws_r, next_row, money_cols=RESUMEN_MONEY, pct_cols=RESUMEN_PCT)
            next_row += 1

    # Helper: suma GA4 atribuido a una lista de campañas (igual lógica que el email)
    def ga4_sum_for_campaigns(campaigns_list):
        matched = set()
        totals  = {"sesiones": 0, "add_to_cart": 0, "begin_checkout": 0, "purchase": 0, "revenue": 0.0}
        for r in campaigns_list:
            g = ga4_by_campaign(r["campana"])
            if not g:
                continue
            key = id(g)
            if key in matched:
                continue
            matched.add(key)
            totals["sesiones"]       += g.get("sesiones", 0)
            totals["add_to_cart"]    += g.get("add_to_cart", 0)
            totals["begin_checkout"] += g.get("begin_checkout", 0)
            totals["purchase"]       += g.get("purchase", 0)
            totals["revenue"]        += g.get("revenue", 0.0)
        return totals

    # ── Hoja Meta Ads ─────────────────────────────────────────────────────────
    if update_sheets:
        # Col: 1=Fecha 2=Inv 3=Alc 4=Impr 5=CPM 6=Frec 7=Clk 8=CPC
        #      9=CTR 10=ComprasMeta 11=CPAMeta 12=ValorMeta
        #      13=Ses 14=Cart 15=Chk 16=ComprasGA4 17=Ing 18=ROAS 19=Conv%
        META_MONEY = (2, 5, 8, 11, 12, 17)   # Inversión, CPM, CPC, CPA Meta, Valor Meta, Ingresos
        META_PCT   = (9, 19)                  # CTR, Conv %

        META_HEADERS = [
            "Fecha", "Inversión", "Alcance", "Impresiones",
            "CPM", "Frecuencia", "Clicks", "CPC", "CTR",
            "Compras Meta", "CPA Meta", "Valor Meta",
            "Sesiones", "Add Cart", "Checkout", "Compras GA4", "Ingresos", "ROAS", "Conv %",
        ]

        # Si la hoja ya existe con estructura anterior, reconstruirla
        if "Meta Ads" in wb.sheetnames:
            ws_check = wb["Meta Ads"]
            col9_val  = ws_check.cell(row=3, column=9).value
            col12_val = ws_check.cell(row=3, column=12).value
            # Reconstruir si CTR no está en col 9 O Valor Meta no está en col 12
            if col9_val != "CTR" or col12_val != "Valor Meta":
                del wb["Meta Ads"]

        ws_m = ensure_sheet(wb, "Meta Ads", META_HEADERS)

        # Reposicionar Meta Ads justo después de Resumen (siempre)
        if "Resumen" in wb.sheetnames:
            resumen_idx = wb.sheetnames.index("Resumen")
            meta_idx    = wb.sheetnames.index("Meta Ads")
            wb.move_sheet("Meta Ads", offset=resumen_idx + 1 - meta_idx)

        # Totales del período (GA4 por atribución de campaña)
        meta_ga4   = ga4_sum_for_campaigns(meta)
        meta_total = defaultdict(float)
        for r in meta_daily:
            meta_total["inversion"]    += r["inversion"]
            meta_total["alcance"]      += r["alcance"]
            meta_total["impresiones"]  += r["impresiones"]
            meta_total["clicks"]       += r["clicks"]
            meta_total["compras_meta"] += r.get("compras_meta", 0)
            meta_total["valor_meta"]   += r.get("valor_meta", 0.0)

        t_inv        = meta_total["inversion"]
        t_impr       = meta_total["impresiones"]
        t_clk        = meta_total["clicks"]
        t_alc        = meta_total["alcance"]
        t_cmp_meta   = int(meta_total["compras_meta"])
        t_valor_meta = round(meta_total["valor_meta"], 2)
        t_ses        = meta_ga4["sesiones"]
        t_pur        = meta_ga4["purchase"]
        t_rev        = meta_ga4["revenue"]

        # ── Limpiar filas de consolidado duplicadas (filas con "→" fuera de fila 4)
        for rr in reversed(range(5, ws_m.max_row + 1)):
            if "→" in str(ws_m.cell(row=rr, column=1).value or ""):
                ws_m.delete_rows(rr)

        # ── Fila 4: consolidado del período (sobreescribir o crear) ───────────
        meta_summ_data = [
            f"{start_date} → {end_date}",
            t_inv,
            int(t_alc)  or None,
            int(t_impr) or None,
            round(t_inv / t_impr * 1000, 2) if t_impr > 0 else None,
            round(t_impr / t_alc,        2) if t_alc  > 0 else None,
            int(t_clk) or None,
            round(t_inv / t_clk,         2) if t_clk  > 0 else None,
            round(t_clk / t_impr * 100,  2) if t_impr > 0 else None,   # CTR (col 9)
            t_cmp_meta or None,                                          # Compras Meta (col 10)
            round(t_inv / t_cmp_meta,    2) if t_cmp_meta > 0 else None,  # CPA Meta (col 11)
            t_valor_meta or None,                                         # Valor Meta (col 12)
            t_ses  or None,
            meta_ga4["add_to_cart"]    or None,
            meta_ga4["begin_checkout"] or None,
            t_pur  or None,
            t_rev  or None,
            round(t_rev / t_inv,         2) if t_inv  > 0 and t_rev > 0 else None,
            round(t_pur / t_ses * 100,   2) if t_ses  > 0 else None,
        ]
        # Buscar fila existente con "→" en col A (debería ser fila 4)
        summ_row_m = next(
            (rr for rr in range(4, min(6, ws_m.max_row + 1))
             if "→" in str(ws_m.cell(row=rr, column=1).value or "")),
            None
        )
        if summ_row_m:
            for col, val in enumerate(meta_summ_data, 1):
                ws_m.cell(row=summ_row_m, column=col).value = val
            fmt_row(ws_m, summ_row_m, money_cols=META_MONEY, pct_cols=META_PCT)
        else:
            append_summary_row(ws_m, meta_summ_data, money_cols=META_MONEY, pct_cols=META_PCT)

        # Filas diarias (con dedup)
        meta_by_day = defaultdict(lambda: {
            "inversion": 0.0, "alcance": 0, "impresiones": 0,
            "clicks": 0, "compras_meta": 0, "valor_meta": 0.0,
        })
        for r in meta_daily:
            d = meta_by_day[r["fecha"]]
            d["inversion"]    += r["inversion"]
            d["alcance"]      += r["alcance"]
            d["impresiones"]  += r["impresiones"]
            d["clicks"]       += r["clicks"]
            d["compras_meta"] += r.get("compras_meta", 0)
            d["valor_meta"]   += r.get("valor_meta", 0.0)

        existing_m = _sheet_dates(ws_m)
        for fecha in sorted(meta_by_day):
            if fecha in existing_m:
                continue
            d    = meta_by_day[fecha]
            inv  = d["inversion"]
            impr = d["impresiones"]
            clk  = d["clicks"]
            alc  = d["alcance"]
            cmp  = int(d["compras_meta"])
            val  = round(d["valor_meta"], 2)
            ws_m.append([
                fecha, inv, alc or None, impr or None,
                round(inv / impr * 1000, 2) if impr > 0 else None,
                round(impr / alc,        2) if alc  > 0 else None,
                clk or None,
                round(inv / clk,         2) if clk  > 0 else None,
                round(clk / impr * 100,  2) if impr > 0 else None,   # CTR (col 9)
                cmp or None,                                           # Compras Meta (col 10)
                round(inv / cmp,         2) if cmp  > 0 else None,   # CPA Meta (col 11)
                val or None,                                           # Valor Meta (col 12)
                None, None, None, None, None, None, None,
            ])
            fmt_row(ws_m, ws_m.max_row, money_cols=META_MONEY, pct_cols=META_PCT)

    # ── Hoja Google Ads ───────────────────────────────────────────────────────
    if update_sheets:
        GOOGLE_MONEY = (2, 4, 6, 12)
        GOOGLE_PCT   = (7, 14)

        GOOGLE_HEADERS = [
            "Fecha", "Inversión", "Impresiones", "CPM", "Clicks", "CPC", "CTR",
            "Sesiones", "Add Cart", "Checkout", "Compras", "Ingresos", "ROAS", "Conv %",
        ]
        ws_g = ensure_sheet(wb, "Google Ads", GOOGLE_HEADERS)

        google_ga4   = ga4_sum_for_campaigns(google)
        google_total = defaultdict(float)
        for r in google_daily:
            google_total["inversion"]   += r["inversion"]
            google_total["impresiones"] += r["impresiones"]
            google_total["clicks"]      += r["clicks"]

        g_inv  = google_total["inversion"]
        g_impr = google_total["impresiones"]
        g_clk  = google_total["clicks"]
        g_ses  = google_ga4["sesiones"]
        g_pur  = google_ga4["purchase"]
        g_rev  = google_ga4["revenue"]

        # ── Limpiar filas de consolidado duplicadas (filas con "→" fuera de fila 4)
        for rr in reversed(range(5, ws_g.max_row + 1)):
            if "→" in str(ws_g.cell(row=rr, column=1).value or ""):
                ws_g.delete_rows(rr)

        # ── Fila 4: consolidado del período (sobreescribir o crear) ───────────
        google_summ_data = [
            f"{start_date} → {end_date}",
            g_inv,
            int(g_impr) or None,
            round(g_inv / g_impr * 1000, 2) if g_impr > 0 else None,
            int(g_clk) or None,
            round(g_inv / g_clk,         2) if g_clk  > 0 else None,
            round(g_clk / g_impr * 100,  2) if g_impr > 0 else None,
            g_ses  or None,
            google_ga4["add_to_cart"]    or None,
            google_ga4["begin_checkout"] or None,
            g_pur  or None,
            g_rev  or None,
            round(g_rev / g_inv,         2) if g_inv  > 0 and g_rev > 0 else None,
            round(g_pur / g_ses * 100,   2) if g_ses  > 0 else None,
        ]
        summ_row_g = next(
            (rr for rr in range(4, min(6, ws_g.max_row + 1))
             if "→" in str(ws_g.cell(row=rr, column=1).value or "")),
            None
        )
        if summ_row_g:
            for col, val in enumerate(google_summ_data, 1):
                ws_g.cell(row=summ_row_g, column=col).value = val
            fmt_row(ws_g, summ_row_g, money_cols=GOOGLE_MONEY, pct_cols=GOOGLE_PCT)
        else:
            append_summary_row(ws_g, google_summ_data, money_cols=GOOGLE_MONEY, pct_cols=GOOGLE_PCT)

        google_by_day = defaultdict(lambda: {"inversion": 0.0, "impresiones": 0, "clicks": 0})
        for r in google_daily:
            d = google_by_day[r["fecha"]]
            d["inversion"]   += r["inversion"]
            d["impresiones"] += r["impresiones"]
            d["clicks"]      += r["clicks"]

        existing_g = _sheet_dates(ws_g)
        for fecha in sorted(google_by_day):
            if fecha in existing_g:
                continue
            d    = google_by_day[fecha]
            inv  = d["inversion"]
            impr = d["impresiones"]
            clk  = d["clicks"]
            ws_g.append([
                fecha, inv, impr or None,
                round(inv / impr * 1000, 2) if impr > 0 else None,
                clk or None,
                round(inv / clk,         2) if clk  > 0 else None,
                round(clk / impr * 100,  2) if impr > 0 else None,
                None, None, None, None, None, None, None,  # GA4 solo en resumen
            ])
            fmt_row(ws_g, ws_g.max_row, money_cols=GOOGLE_MONEY, pct_cols=GOOGLE_PCT)

    # ── Hoja "reporte del día" (una por envío, igual que el mail) ────────────
    sheet_name = datetime.strptime(end_date, "%Y-%m-%d").strftime("%d-%m-%Y")
    if add_status_tab and sheet_name not in wb.sheetnames:
        RPT_HEADERS = [
            "Campaña", "Inversión", "Alcance", "Impr.", "CPM", "Frec.",
            "Clicks", "CPC", "CTR", "Usuarios GA4", "Sesiones",
            "Add Cart", "Checkout", "Compras", "Ingresos", "ROAS", "Conv.%",
            "Compras Meta", "CPA Meta", "Valor Meta",
        ]
        RPT_MONEY = (2, 5, 8, 15, 19, 20)   # Inversión, CPM, CPC, Ingresos, CPA Meta, Valor Meta
        RPT_PCT   = (9, 17)                  # CTR, Conv.%
        ncols_rpt = len(RPT_HEADERS)

        ws_rpt = wb.create_sheet(sheet_name)

        # Fila 1: encabezado azul con logo
        ws_rpt.row_dimensions[1].height = 48
        last_rpt_col = get_column_letter(ncols_rpt)
        ws_rpt.merge_cells(f"A1:{last_rpt_col}1")
        hdr = ws_rpt["A1"]
        hdr.fill      = PatternFill("solid", fgColor=BLUE)
        hdr.font      = Font(bold=True, color="FFFFFF", size=14, name=HF)
        hdr.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        hdr.value     = f"MOSCA HNOS.  ·  Reporte {_fmt(start_date)} → {_fmt(end_date)}  "
        if os.path.exists(logo_path):
            img_r = XLImage(logo_path); img_r.width = 44; img_r.height = 44
            ws_rpt.add_image(img_r, "A1")

        # Fila 2: separador
        ws_rpt.row_dimensions[2].height = 6
        for col in range(1, ncols_rpt + 1):
            ws_rpt.cell(row=2, column=col).fill = PatternFill("solid", fgColor=BLUE)

        # Fila 3: cabeceras
        ws_rpt.row_dimensions[3].height = 20
        for col, h in enumerate(RPT_HEADERS, 1):
            c = ws_rpt.cell(row=3, column=col, value=h)
            c.font      = COL_FONT
            c.fill      = COL_FILL
            c.alignment = COL_ALIGN
            ws_rpt.column_dimensions[get_column_letter(col)].width = max(14, len(h) + 2)
        ws_rpt.column_dimensions["A"].width = 42  # columna Campaña más ancha

        def _plat_totals(rows):
            inv     = sum(r["inversion"] for r in rows)
            alcance = sum(r["alcance"] for r in rows if isinstance(r.get("alcance"), (int, float)))
            impr    = sum(r["impresiones"] for r in rows if isinstance(r.get("impresiones"), (int, float)))
            clicks  = sum(r["clicks"] for r in rows if isinstance(r.get("clicks"), (int, float)))
            cpc     = (inv / clicks)      if clicks > 0 else None
            cpm     = (inv / impr * 1000) if impr   > 0 else None
            g_u = g_ses = g_cart = g_chk = g_pur = 0; g_rev = 0.0
            for r in rows:
                gd = ga4_by_campaign(r["campana"])
                g_u   += gd.get("usuarios",       0)
                g_ses += gd.get("sesiones",        0)
                g_cart+= gd.get("add_to_cart",     0)
                g_chk += gd.get("begin_checkout",  0)
                g_pur += gd.get("purchase",        0)
                g_rev += gd.get("revenue",         0.0)
            roas = (g_rev / inv)        if inv   > 0 and g_rev > 0 else None
            conv = (g_pur / g_ses * 100) if g_ses > 0 else None
            freq = None
            # Meta-native totals (solo aplica para filas Meta)
            cmp_meta = int(sum(r.get("compras_meta", 0) for r in rows))
            cpa_meta = round(inv / cmp_meta, 2) if cmp_meta > 0 else None
            val_meta = round(sum(r.get("valor_meta", 0.0) for r in rows), 2)
            return [inv, alcance or None, impr or None, cpm, freq, clicks or None,
                    cpc, None, g_u or None, g_ses or None,
                    g_cart or None, g_chk or None, g_pur or None,
                    g_rev or None, roas, conv,
                    cmp_meta or None, cpa_meta, val_meta or None]

        # Fila 4: totales globales
        all_rpt = meta + google + tiktok
        tot     = _plat_totals(all_rpt)
        SUMM_FILL_RPT = PatternFill("solid", fgColor="D6E4F0")
        SUMM_FONT_RPT = Font(bold=True, size=10, name=HF)
        total_row = ["TOTAL PERÍODO"] + tot
        ws_rpt.append(total_row)
        for col in range(1, ncols_rpt + 1):
            c = ws_rpt.cell(row=ws_rpt.max_row, column=col)
            c.fill = SUMM_FILL_RPT
            c.font = SUMM_FONT_RPT
            c.alignment = Alignment(horizontal="right" if col > 1 else "left", vertical="center")
        fmt_row(ws_rpt, ws_rpt.max_row, money_cols=RPT_MONEY, pct_cols=RPT_PCT)

        # Helper: fila de plataforma (encabezado coloreado) + filas de campaña
        PLAT_HEX = {"Meta": "1877F2", "Google": "4285F4", "TikTok": "010101"}

        def _add_platform_rows(ws, rows, label):
            if not rows:
                return
            color = PLAT_HEX.get(label, BLUE)
            plat_fill = PatternFill("solid", fgColor=color)
            plat_font = Font(bold=True, color="FFFFFF", size=10, name=HF)

            # Totales de plataforma como encabezado
            pt = _plat_totals(rows)
            hdr_row = [f"{label.upper()} ADS  ·  {len(rows)} campañas"] + pt
            ws.append(hdr_row)
            ridx = ws.max_row
            for col in range(1, ncols_rpt + 1):
                c = ws.cell(row=ridx, column=col)
                c.fill      = plat_fill
                c.font      = plat_font
                c.alignment = Alignment(horizontal="right" if col > 1 else "left", vertical="center")
            fmt_row(ws, ridx, money_cols=RPT_MONEY, pct_cols=RPT_PCT)

            # Filas de campaña
            for i, r in enumerate(rows):
                gd   = ga4_by_campaign(r["campana"])
                inv  = r["inversion"]
                alc  = r.get("alcance")
                impr = r.get("impresiones")
                clk  = r.get("clicks")
                rev  = gd.get("revenue", 0.0)
                pur  = gd.get("purchase", 0)
                ses  = gd.get("sesiones", 0)
                roas = (rev / inv)        if inv > 0 and rev > 0 else None
                conv = (pur / ses * 100)  if ses > 0 else None
                freq = r.get("frecuencia")
                ctr  = r.get("ctr")
                cmp_meta_r = r.get("compras_meta") or None
                cpa_meta_r = r.get("cpa_meta") or None
                val_meta_r = r.get("valor_meta") or None
                row_data = [
                    r["campana"],
                    inv,
                    alc  if isinstance(alc,  (int, float)) else None,
                    impr if isinstance(impr, (int, float)) else None,
                    r.get("cpm"),
                    freq if isinstance(freq, (int, float)) else None,
                    clk  if isinstance(clk,  (int, float)) else None,
                    r.get("cpc"),
                    ctr  if isinstance(ctr,  (int, float)) else None,
                    gd.get("usuarios")       or None,
                    ses  or None,
                    gd.get("add_to_cart")    or None,
                    gd.get("begin_checkout") or None,
                    pur  or None,
                    rev  or None,
                    roas,
                    conv,
                    cmp_meta_r,
                    cpa_meta_r,
                    val_meta_r,
                ]
                ws.append(row_data)
                ridx = ws.max_row
                row_fill = PatternFill("solid", fgColor="F9F9F9" if i % 2 == 0 else "FFFFFF")
                for col in range(1, ncols_rpt + 1):
                    c = ws.cell(row=ridx, column=col)
                    c.fill      = row_fill
                    c.font      = Font(size=10, name=HF)
                    c.alignment = Alignment(horizontal="right" if col > 1 else "left", vertical="center")
                fmt_row(ws, ridx, money_cols=RPT_MONEY, pct_cols=RPT_PCT)

        _add_platform_rows(ws_rpt, meta,    "Meta")
        _add_platform_rows(ws_rpt, google,  "Google")
        _add_platform_rows(ws_rpt, tiktok,  "TikTok")

        # ── 6 gráficos de torta Meta vs Google (desde col X) ─────────────────
        from openpyxl.chart.label import DataLabelList as _DLL

        meta_pt   = _plat_totals(meta)
        google_pt = _plat_totals(google)
        # índices de _plat_totals: inv(0), impr(2), clicks(5), g_ses(9), g_pur(12), g_rev(13)
        chart_metrics = [
            ("Inversión",   meta_pt[0],  google_pt[0]),
            ("Impresiones", meta_pt[2],  google_pt[2]),
            ("Clicks",      meta_pt[5],  google_pt[5]),
            ("Sesiones",    meta_pt[9],  google_pt[9]),
            ("Compras",     meta_pt[12], google_pt[12]),
            ("Ingresos",    meta_pt[13], google_pt[13]),
        ]

        # Tablas de datos en col Z(26)/AA(27), costos en AB(28)/AC(29)
        TBL_COL_L  = 26   # Z  — etiquetas pie
        TBL_COL_V  = 27   # AA — valores pie
        # col AB (28) queda vacía como separador
        COST_COL_L = 29   # AC — etiquetas costo
        COST_COL_V = 30   # AD — valores costo
        tbl_rows   = [4, 8, 12, 16, 20, 24]  # fila de inicio de cada tabla

        # Costos unitarios por métrica (costo_meta, costo_google, etiqueta_columna)
        m_inv = meta_pt[0] or 0
        g_inv = google_pt[0] or 0
        cost_by_row = {
            8:  ("CPM",
                 round(m_inv / meta_pt[2]   * 1000, 2) if meta_pt[2]   else None,
                 round(g_inv / google_pt[2] * 1000, 2) if google_pt[2] else None),
            12: ("CPC",
                 round(m_inv / meta_pt[5],   2) if meta_pt[5]   else None,
                 round(g_inv / google_pt[5], 2) if google_pt[5] else None),
            16: ("Costo/Sesión",
                 round(m_inv / meta_pt[9],   2) if meta_pt[9]   else None,
                 round(g_inv / google_pt[9], 2) if google_pt[9] else None),
            20: ("CPA",
                 round(m_inv / meta_pt[12],   2) if meta_pt[12]   else None,
                 round(g_inv / google_pt[12], 2) if google_pt[12] else None),
        }

        # Anchors: grilla 3×2 — AF(32), AM(39), AU(47)
        COL_ANCHORS = [32, 39, 47]
        ROW_ANCHORS = [4, 22]
        chart_anchors = [
            get_column_letter(c) + str(r)
            for r in ROW_ANCHORS for c in COL_ANCHORS
        ]

        for idx, (label, meta_val, google_val) in enumerate(chart_metrics):
            tbl_r = tbl_rows[idx]

            # Escribir tabla de datos (Z:AA)
            ws_rpt.cell(row=tbl_r,     column=TBL_COL_L, value="Plataforma").font = COL_FONT
            ws_rpt.cell(row=tbl_r,     column=TBL_COL_L).fill = COL_FILL
            ws_rpt.cell(row=tbl_r,     column=TBL_COL_V, value=label).font = COL_FONT
            ws_rpt.cell(row=tbl_r,     column=TBL_COL_V).fill = COL_FILL
            ws_rpt.cell(row=tbl_r + 1, column=TBL_COL_L, value="Meta Ads")
            ws_rpt.cell(row=tbl_r + 2, column=TBL_COL_L, value="Google Ads")
            c_mv = ws_rpt.cell(row=tbl_r + 1, column=TBL_COL_V, value=meta_val   or 0)
            c_gv = ws_rpt.cell(row=tbl_r + 2, column=TBL_COL_V, value=google_val or 0)
            # Formato monetario para Inversión e Ingresos
            if label in ("Inversión", "Ingresos"):
                c_mv.number_format = USD
                c_gv.number_format = USD

            # Tabla de costos (AB:AC) para métricas con costo unitario
            if tbl_r in cost_by_row:
                cost_label, cost_meta, cost_google = cost_by_row[tbl_r]
                ws_rpt.cell(row=tbl_r,     column=COST_COL_L, value="Plataforma").font  = COL_FONT
                ws_rpt.cell(row=tbl_r,     column=COST_COL_L).fill = COL_FILL
                ws_rpt.cell(row=tbl_r,     column=COST_COL_V, value=cost_label).font   = COL_FONT
                ws_rpt.cell(row=tbl_r,     column=COST_COL_V).fill = COL_FILL
                ws_rpt.cell(row=tbl_r + 1, column=COST_COL_L, value="Meta Ads")
                ws_rpt.cell(row=tbl_r + 2, column=COST_COL_L, value="Google Ads")
                cm = ws_rpt.cell(row=tbl_r + 1, column=COST_COL_V, value=cost_meta   or 0)
                cg = ws_rpt.cell(row=tbl_r + 2, column=COST_COL_V, value=cost_google or 0)
                cm.number_format = USD
                cg.number_format = USD

            # Crear gráfico de torta
            pc = PieChart()
            pc.title  = label
            pc.style  = 2
            pc.width  = 8
            pc.height = 7

            pc_data = Reference(ws_rpt, min_col=TBL_COL_V, min_row=tbl_r,     max_row=tbl_r + 2)
            pc_cats = Reference(ws_rpt, min_col=TBL_COL_L, min_row=tbl_r + 1, max_row=tbl_r + 2)
            pc.add_data(pc_data, titles_from_data=True)
            pc.set_categories(pc_cats)

            dL = _DLL()
            dL.showPercent   = True
            dL.showVal       = False
            dL.showCatName   = True
            dL.showSerName   = False
            dL.showLegendKey = False
            pc.series[0].dLbls = dL

            pc.anchor = chart_anchors[idx]
            ws_rpt.add_chart(pc)

        ws_rpt.column_dimensions[get_column_letter(TBL_COL_L)].width  = 14
        ws_rpt.column_dimensions[get_column_letter(TBL_COL_V)].width  = 14
        ws_rpt.column_dimensions[get_column_letter(COST_COL_L)].width = 14
        ws_rpt.column_dimensions[get_column_letter(COST_COL_V)].width = 14

        # Mover la nueva pestaña al final (después de las otras)
        wb.move_sheet(sheet_name, offset=len(wb.sheetnames) - 1)

    wb.save(excel_path)
    print(f"✓ Excel evolutivo guardado: {excel_path}")
    return excel_path


# ── GOOGLE DRIVE ──────────────────────────────────────────────────────────────

DRIVE_FOLDER_ID = "1RsFuofmXTJdGgHxQPdzdVZ8pitwvlsr8"

def upload_to_drive(local_path):
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload

    creds = service_account.Credentials.from_service_account_file(
        GA4_KEY_FILE,
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    service = build("drive", "v3", credentials=creds)

    filename = os.path.basename(local_path)
    mime     = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # Buscar si ya existe un archivo con ese nombre en la carpeta para actualizarlo
    existing = service.files().list(
        q=f"name='{filename}' and '{DRIVE_FOLDER_ID}' in parents and trashed=false",
        fields="files(id)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute().get("files", [])

    media = MediaFileUpload(local_path, mimetype=mime, resumable=False)

    if existing:
        file_id = existing[0]["id"]
        service.files().update(
            fileId=file_id, media_body=media, supportsAllDrives=True,
        ).execute()
        print(f"✓ Excel actualizado en Drive: {filename} (id: {file_id})")
    else:
        meta = {"name": filename, "parents": [DRIVE_FOLDER_ID]}
        f    = service.files().create(
            body=meta, media_body=media, fields="id", supportsAllDrives=True,
        ).execute()
        print(f"✓ Excel subido a Drive: {filename} (id: {f['id']})")


# ── ENVIAR MAIL ───────────────────────────────────────────────────────────────

def send_email(subject, html_body):
    from email.mime.image import MIMEImage
    recipients = [r.strip() for r in EMAIL_RECIPIENT.split(",")]

    msg            = MIMEMultipart("related")
    msg["Subject"] = subject
    msg["From"]    = EMAIL_SENDER
    msg["To"]      = ", ".join(recipients)

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(html_body, "html", "utf-8"))
    msg.attach(alt)

    # Adjuntar logo como imagen inline
    logo_path = os.path.join(os.path.dirname(__file__), "logo-mosca.png")
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            img = MIMEImage(f.read())
            img.add_header("Content-ID", "<logo_mosca>")
            img.add_header("Content-Disposition", "inline", filename="logo-mosca.png")
            msg.attach(img)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(EMAIL_SENDER, EMAIL_APP_PASSWORD)
        server.sendmail(EMAIL_SENDER, recipients, msg.as_string())

    print(f"✓ Email enviado a {', '.join(recipients)}")


# ── ANÁLISIS IA DE CAMPAÑAS ───────────────────────────────────────────────────

def generate_campaign_analysis(start_date, end_date, meta, google, ga4, excel_path):
    """Genera análisis IA de campañas y devuelve el texto. Requiere ANTHROPIC_API_KEY."""
    import anthropic as _anthropic
    import openpyxl

    # GA4 lookup por nombre de campaña
    def _ga4_lookup(campana):
        name = campana.lower()
        if name in ga4:
            return ga4[name]
        for k, v in ga4.items():
            if name in k or k in name:
                return v
        return {}

    def _fmt_campaigns(rows):
        lines = []
        for r in rows:
            gd   = _ga4_lookup(r["campana"])
            inv  = r.get("inversion") or 0
            ses  = gd.get("sesiones", 0)
            pur  = gd.get("purchase", 0)
            rev  = gd.get("revenue",  0.0)
            roas = round(rev / inv, 2)       if inv > 0 and rev > 0 else "N/D"
            conv = round(pur / ses * 100, 2) if ses > 0              else "N/D"
            lines.append(
                f"  - {r['campana']}: Inv=${inv:.0f} | "
                f"Impr={r.get('impresiones', 'N/D')} | Clicks={r.get('clicks', 'N/D')} | "
                f"Ses={ses} | Compras={pur} | Ingresos=${rev:.0f} | "
                f"ROAS={roas} | Tasa cierre={conv}%"
            )
        return "\n".join(lines) if lines else "  (sin datos)"

    current_block = (
        f"=== PERÍODO ACTUAL: {start_date} → {end_date} ===\n\n"
        f"META ADS:\n{_fmt_campaigns(meta)}\n\n"
        f"GOOGLE ADS:\n{_fmt_campaigns(google)}"
    )

    # Datos del status tab anterior (desde el Excel)
    prev_block = ""
    try:
        wb         = openpyxl.load_workbook(excel_path, read_only=True)
        curr_sn    = datetime.strptime(end_date, "%Y-%m-%d").strftime("%d-%m-%Y")
        date_sheets = []
        for sn in wb.sheetnames:
            try:
                dt = datetime.strptime(sn, "%d-%m-%Y")
                if sn != curr_sn:
                    date_sheets.append((dt, sn))
            except ValueError:
                pass
        if date_sheets:
            prev_sn    = max(date_sheets, key=lambda x: x[0])[1]
            ws_prev    = wb[prev_sn]
            prev_lines = []
            for rr in range(4, ws_prev.max_row + 1):
                campana = ws_prev.cell(row=rr, column=1).value
                if not campana or not str(campana).strip():
                    continue
                inv  = ws_prev.cell(row=rr, column=2).value  or 0
                ses  = ws_prev.cell(row=rr, column=11).value or 0
                pur  = ws_prev.cell(row=rr, column=14).value or 0
                rev  = ws_prev.cell(row=rr, column=15).value or 0
                roas = ws_prev.cell(row=rr, column=16).value
                conv = ws_prev.cell(row=rr, column=17).value
                prev_lines.append(
                    f"  - {campana}: Inv=${inv:.0f} | Ses={ses} | "
                    f"Compras={pur} | Ingresos=${rev:.0f} | ROAS={roas} | Tasa cierre={conv}"
                )
            if prev_lines:
                prev_block = (
                    f"\n=== PERÍODO ANTERIOR (status {prev_sn}) ===\n"
                    + "\n".join(prev_lines)
                )
        wb.close()
    except Exception:
        pass

    prompt = f"""Sos un especialista senior en performance marketing digital para e-commerce. Analizá los datos de campañas de Mosca Hnos. (tienda de juguetes, librería y artículos de primera infancia en Uruguay) y producir un análisis estratégico.

CONTEXTO DEL FUNNEL:
Las campañas cumplen roles distintos en el funnel de conversión:
- Campañas de ALCANCE / TRÁFICO (Meta): parte alta del funnel — generan awareness y audiencias para retargeting. No se espera conversión directa pero sí alimentan las etapas siguientes.
- Campañas de REMARKETING (Meta): parte media — impactan usuarios que ya visitaron el sitio. Se espera mayor tasa de conversión que alcance pero menor que brand search.
- Campañas SEARCH BRAND (Google): parte baja — capturan demanda generada por alcance. Alta intención de compra. Son el indicador más directo de la salud de la marca.
- Campañas SEARCH GENÉRICO (Google): parte media-baja — capturan demanda de categoría. La conversión depende del precio, catálogo y UX.
- Campañas PMAX / SHOPPING (Google): automatizadas — cubren múltiples canales, complementan a Brand y Search Genérico.

{current_block}
{prev_block}

Generá el análisis con exactamente estas secciones (títulos en mayúscula):

1. RESUMEN EJECUTIVO
Inversión total, ingresos totales, ROAS global, tasa de cierre general. Lectura rápida en 4-5 oraciones.

2. LECTURA DEL FUNNEL
¿Las campañas de alcance están alimentando el brand search? ¿El remarketing está convirtiendo? ¿Hay cuellos de botella?

3. META VS GOOGLE
Eficiencia comparada: ROAS, costo por sesión, tasa de cierre. Rol de cada plataforma en el mix.

4. CAMPAÑAS DESTACADAS
Las que mejor performaron. Citá nombre, ROAS y tasa de cierre.

5. CAMPAÑAS A REVISAR
Problema concreto de cada una y acción a tomar.

6. EVOLUCIÓN VS PERÍODO ANTERIOR
Solo si hay datos anteriores. Qué cambió en el funnel.

7. RECOMENDACIONES
4 acciones priorizadas para mejorar la tasa de cierre. Cada una: qué hacer, en qué campaña/canal, resultado esperado.

Texto plano, sin asteriscos ni markdown. Usá los nombres exactos de las campañas. Citá números específicos."""

    client   = _anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    response = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=4000,
        messages=[{"role": "user", "content": prompt}],
    )
    return response.content[0].text


def build_notas_email_html(start_date, end_date, analysis_text):
    """Genera el HTML del mail con el análisis IA."""
    BLUE = "003B73"
    HF   = "Arial"

    # Convertir secciones numeradas a HTML
    paragraphs = []
    for line in analysis_text.split("\n"):
        stripped = line.strip()
        if not stripped:
            continue
        # Detectar títulos de sección (ej. "1. RESUMEN EJECUTIVO")
        import re
        if re.match(r"^\d+\.\s+[A-ZÁÉÍÓÚÑ\s]+$", stripped):
            paragraphs.append(
                f'<h3 style="color:#{BLUE};font-family:{HF},sans-serif;'
                f'margin:20px 0 6px 0;font-size:13px;border-bottom:1px solid #ddd;'
                f'padding-bottom:4px;">{stripped}</h3>'
            )
        else:
            paragraphs.append(
                f'<p style="font-family:{HF},sans-serif;font-size:13px;'
                f'color:#222;margin:4px 0;line-height:1.6;">{stripped}</p>'
            )

    body_html = "\n".join(paragraphs)
    fecha_fmt = f"{_fmt(start_date)} → {_fmt(end_date)}"

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f4f6f8;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f6f8;">
  <tr><td align="center" style="padding:30px 10px;">
    <table width="680" cellpadding="0" cellspacing="0"
           style="background:#ffffff;border-radius:8px;overflow:hidden;
                  box-shadow:0 2px 8px rgba(0,0,0,0.08);">

      <!-- Header -->
      <tr>
        <td style="background:#{BLUE};padding:20px 30px;">
          <table width="100%" cellpadding="0" cellspacing="0">
            <tr>
              <td>
                <img src="cid:logo_mosca" width="44" height="44"
                     style="vertical-align:middle;margin-right:12px;">
                <span style="font-family:{HF},sans-serif;font-size:20px;
                             font-weight:bold;color:#ffffff;vertical-align:middle;">
                  MOSCA HNOS.
                </span>
              </td>
              <td align="right">
                <span style="font-family:{HF},sans-serif;font-size:12px;color:#cce0f5;">
                  Análisis de Campañas · {fecha_fmt}
                </span>
              </td>
            </tr>
          </table>
        </td>
      </tr>

      <!-- Body -->
      <tr>
        <td style="padding:30px 30px 20px 30px;">
          {body_html}
        </td>
      </tr>

      <!-- Footer -->
      <tr>
        <td style="background:#f4f6f8;padding:14px 30px;border-top:1px solid #e0e0e0;">
          <p style="font-family:{HF},sans-serif;font-size:11px;color:#888;margin:0;">
            Análisis generado automáticamente por IA · Cardinal
          </p>
        </td>
      </tr>

    </table>
  </td></tr>
</table>
</body>
</html>"""


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    import time, traceback

    # ── Detectar modo de ejecución ────────────────────────────────────────────
    # --daily  → solo actualiza Resumen/Meta/Google con datos de ayer (sin email)
    # 2 args   → manual con fechas explícitas (email + status tab + hojas)
    # sin args → programado lunes/viernes usando fecha actual (email + status tab, sin tocar las 3 hojas)

    args      = sys.argv[1:]
    today_str = datetime.today().strftime("%Y-%m-%d")
    yesterday = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")

    if args == ["--daily"]:
        daily_run  = True
        manual_run = False
        start_date, _ = get_date_range()   # inicio del período actual
        end_date   = yesterday
        print(f"\n[DAILY] Actualizando Excel con período {start_date} → {end_date}")
    elif len(args) == 2:
        daily_run  = False
        manual_run = True
        start_date, end_date = args
        print(f"\n[MANUAL] Período: {start_date} → {end_date}")
    else:
        daily_run  = False
        manual_run = False
        start_date, _ = get_date_range()
        end_date = today_str
        print(f"\n[SEMANAL] Período: {start_date} → {end_date}")
    # ── Reglas de fecha final ──────────────────────────────────────────────────
    # daily: siempre día cerrado (ayer)
    # manual: se recorta solo si viene en futuro
    # semanal (sin args): permite fecha actual
    if daily_run and end_date >= today_str:
        end_date = yesterday
        print(f"      → end_date ajustado a {end_date} (día cerrado)")
    elif manual_run and end_date > today_str:
        end_date = today_str
        print(f"      → end_date ajustado a {end_date} (hoy)")

    # ── Obtener datos ─────────────────────────────────────────────────────────
    print("\n[1/4] Meta Ads...")
    try:
        meta = get_meta_campaigns(start_date, end_date)
        print(f"      ✓ {len(meta)} campañas")
    except Exception as e:
        print(f"      ✗ Error: {e}")
        traceback.print_exc()
        meta = []

    print("[2/4] Google Ads...")
    try:
        google = get_google_campaigns(start_date, end_date)
        print(f"      ✓ {len(google)} campañas")
    except Exception as e:
        print(f"      ✗ Error: {e}")
        google = []

    print("[3/4] TikTok Ads...")
    try:
        tiktok = get_tiktok_campaigns(start_date, end_date)
        print(f"      ✓ {len(tiktok)} campañas")
    except Exception as e:
        print(f"      ✗ Error: {e}")
        tiktok = []

    print("[4/4] Google Analytics 4...")
    try:
        ga4 = get_ga4_data(start_date, end_date)
        print(f"      ✓ {len(ga4)} fuentes")
    except Exception as e:
        print(f"      ✗ Error: {e}")
        ga4 = {}

    try:
        ga4_totals = get_ga4_daily_totals(start_date, end_date)
        total_ses = sum(v.get("sesiones", 0) for v in ga4_totals.values())
        print(f"      ✓ Totales sitio: {total_ses:,} sesiones en {len(ga4_totals)} días")
    except Exception as e:
        print(f"      ✗ Error GA4 totales: {e}")
        ga4_totals = {}

    try:
        ga4_by_channel = get_ga4_daily_by_channel(start_date, end_date)
        print(f"      ✓ GA4 por canal: {len(ga4_by_channel)} días")
    except Exception as e:
        print(f"      ✗ Error GA4 por canal: {e}")
        ga4_by_channel = {}

    print("[5/5] Datos diarios para Excel...")
    try:
        time.sleep(10)  # pausa para evitar rate limit de Meta
        meta_daily = get_meta_daily(start_date, end_date)
        print(f"      ✓ Meta: {len(meta_daily)} filas diarias")
    except Exception as e:
        print(f"      ✗ Error Meta daily: {e}")
        meta_daily = []
    try:
        google_daily = get_google_daily(start_date, end_date)
        print(f"      ✓ Google: {len(google_daily)} filas diarias")
    except Exception as e:
        print(f"      ✗ Error Google daily: {e}")
        google_daily = []

    # ── Email (solo en run semanal o manual) ──────────────────────────────────
    if not daily_run:
        subject = f"MOSCA | Reporte Performance Status semanal | {_fmt(start_date)} → {_fmt(end_date)}"
        html    = build_email_html(start_date, end_date, meta, google, tiktok, ga4)
        print("\nEnviando email...")
        send_email(subject, html)

    # ── Excel: determinar qué actualizar ─────────────────────────────────────
    # daily_run  → update_sheets=True,  add_status_tab=False  (solo filas diarias)
    # semanal    → update_sheets=False, add_status_tab=True   (solo pestaña status)
    # manual     → update_sheets=True,  add_status_tab=True   (todo)
    update_sheets  = daily_run or manual_run
    add_status_tab = not daily_run

    # ── YoY para Resumen (mismo período analizado) ───────────────────────────
    ga4_totals_yoy = None
    inv_yoy        = 0.0
    if update_sheets:
        try:
            start_dt  = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt    = datetime.strptime(end_date,   "%Y-%m-%d")
            start_yoy = start_dt.replace(year=start_dt.year - 1).strftime("%Y-%m-%d")
            end_yoy   = end_dt.replace(  year=end_dt.year   - 1).strftime("%Y-%m-%d")
            print(f"\n[YoY] Obteniendo datos {start_yoy} → {end_yoy}...")
            ga4_totals_yoy = get_ga4_daily_totals(start_yoy, end_yoy)
            meta_yoy       = get_meta_campaigns(start_yoy, end_yoy)
            google_yoy     = get_google_campaigns(start_yoy, end_yoy)
            inv_yoy        = sum(r["inversion"] for r in meta_yoy + google_yoy)
            print(f"      ✓ GA4 YoY: {len(ga4_totals_yoy)} días | Inversión YoY: ${inv_yoy:,.2f}")
        except Exception as e:
            print(f"      ✗ Error YoY: {e}")
            traceback.print_exc()

    print("\nActualizando Excel evolutivo...")
    excel_path = None
    try:
        excel_path = update_excel(
            start_date, end_date, meta, google, tiktok, ga4,
            meta_daily, google_daily, ga4_totals, ga4_by_channel,
            ga4_totals_yoy, inv_yoy,
            update_sheets=update_sheets, add_status_tab=add_status_tab,
        )
        upload_to_drive(excel_path)
    except Exception as e:
        print(f"      ✗ Error al guardar/subir Excel: {e}")
        traceback.print_exc()

    # ── Mail de análisis IA (solo lunes/viernes, si hay API key) ─────────────
    if not daily_run and ANTHROPIC_API_KEY and excel_path:
        print("\nGenerando análisis IA de campañas...")
        try:
            analysis = generate_campaign_analysis(
                start_date, end_date, meta, google, ga4, excel_path
            )
            notas_subject = (
                f"MOSCA | Análisis de Campañas | {_fmt(start_date)} → {_fmt(end_date)}"
            )
            notas_html = build_notas_email_html(start_date, end_date, analysis)
            send_email(notas_subject, notas_html)
            print("✓ Mail de análisis IA enviado")
        except Exception as e:
            print(f"      ✗ Error en análisis IA: {e}")
            traceback.print_exc()

    print("✓ Listo.\n")


if __name__ == "__main__":
    main()
